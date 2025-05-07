import datetime
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Union, Tuple
from dataclasses import dataclass, field, asdict


@dataclass
class Transaction:
    """Represents a financial transaction."""
    date: datetime.date
    description: str
    amount: float
    category: str  # 'revenue' or 'expense'
    subcategory: str  # specific type of revenue or expense
    
    def to_dict(self):
        """Convert transaction to dictionary."""
        result = asdict(self)
        result['date'] = self.date.isoformat()
        return result
    
    @classmethod
    def from_dict(cls, data):
        """Create transaction from dictionary."""
        if isinstance(data['date'], str):
            data['date'] = datetime.date.fromisoformat(data['date'])
        return cls(**data)


class IncomeStatement:
    """Manages transactions and generates income statements."""
    
    # Standard categories for revenue and expenses
    REVENUE_CATEGORIES = [
        "sales", "service_revenue", "interest_income", "rental_income", 
        "commission_income", "other_revenue"
    ]
    
    EXPENSE_CATEGORIES = [
        "cost_of_goods_sold", "salaries", "rent", "utilities", 
        "office_supplies", "marketing", "insurance", "depreciation", 
        "interest_expense", "taxes", "other_expenses"
    ]
    
    def __init__(self):
        self.transactions: List[Transaction] = []
        
    def add_transaction(self, 
                      date: Union[str, datetime.date], 
                      description: str, 
                      amount: float, 
                      category: str, 
                      subcategory: str) -> Transaction:
        """
        Add a transaction to the income statement.
        
        Args:
            date: Transaction date (string YYYY-MM-DD or datetime.date)
            description: Description of the transaction
            amount: Transaction amount (positive for revenue, negative for expenses)
            category: Either 'revenue' or 'expense'
            subcategory: Specific type of revenue or expense
        """
        if isinstance(date, str):
            date = datetime.date.fromisoformat(date)
            
        # Validate category
        if category.lower() not in ['revenue', 'expense']:
            raise ValueError("Category must be either 'revenue' or 'expense'")
        
        # Create transaction
        transaction = Transaction(
            date=date,
            description=description,
            amount=amount,
            category=category.lower(),
            subcategory=subcategory.lower()
        )
        
        self.transactions.append(transaction)
        return transaction
    
    def categorize_transaction(self, transaction: Transaction, 
                             new_category: str, new_subcategory: str) -> None:
        """Update transaction categorization."""
        if new_category.lower() not in ['revenue', 'expense']:
            raise ValueError("Category must be either 'revenue' or 'expense'")
            
        transaction.category = new_category.lower()
        transaction.subcategory = new_subcategory.lower()
    
    def get_transactions_by_period(self, start_date=None, end_date=None) -> List[Transaction]:
        """Get transactions within the specified date range."""
        if start_date and isinstance(start_date, str):
            start_date = datetime.date.fromisoformat(start_date)
        
        if end_date and isinstance(end_date, str):
            end_date = datetime.date.fromisoformat(end_date)
        
        filtered = self.transactions
        
        if start_date:
            filtered = [t for t in filtered if t.date >= start_date]
            
        if end_date:
            filtered = [t for t in filtered if t.date <= end_date]
            
        return filtered
    
    def generate_income_statement(self, start_date=None, end_date=None) -> Dict:
        """
        Generate income statement for the specified period.
        
        Returns:
            Dictionary with revenue, expenses, and net income data
        """
        transactions = self.get_transactions_by_period(start_date, end_date)
        
        # Initialize results structure
        result = {
            "period": {
                "start_date": start_date.isoformat() if start_date else "beginning",
                "end_date": end_date.isoformat() if end_date else "present"
            },
            "revenue": {},
            "expenses": {},
            "summary": {
                "total_revenue": 0.0,
                "total_expenses": 0.0,
                "net_income": 0.0
            }
        }
        
        # Categorize and sum transactions
        for transaction in transactions:
            if transaction.category == 'revenue':
                if transaction.subcategory not in result["revenue"]:
                    result["revenue"][transaction.subcategory] = 0.0
                result["revenue"][transaction.subcategory] += transaction.amount
                result["summary"]["total_revenue"] += transaction.amount
            elif transaction.category == 'expense':
                if transaction.subcategory not in result["expenses"]:
                    result["expenses"][transaction.subcategory] = 0.0
                result["expenses"][transaction.subcategory] += transaction.amount
                result["summary"]["total_expenses"] += transaction.amount
        
        # Calculate net income
        result["summary"]["net_income"] = result["summary"]["total_revenue"] - result["summary"]["total_expenses"]
        
        return result
    
    def print_income_statement(self, start_date=None, end_date=None) -> None:
        """Print formatted income statement to console."""
        statement = self.generate_income_statement(start_date, end_date)
        
        print("\n" + "="*60)
        print(f"INCOME STATEMENT")
        print(f"Period: {statement['period']['start_date']} to {statement['period']['end_date']}")
        print("="*60)
        
        print("\nREVENUE:")
        for category, amount in statement['revenue'].items():
            print(f"  {category.replace('_', ' ').title()}: ${amount:.2f}")
        print(f"  {'='*30}")
        print(f"  Total Revenue: ${statement['summary']['total_revenue']:.2f}")
        
        print("\nEXPENSES:")
        for category, amount in statement['expenses'].items():
            print(f"  {category.replace('_', ' ').title()}: ${amount:.2f}")
        print(f"  {'='*30}")
        print(f"  Total Expenses: ${statement['summary']['total_expenses']:.2f}")
        
        print("\nSUMMARY:")
        print(f"  Total Revenue: ${statement['summary']['total_revenue']:.2f}")
        print(f"  Total Expenses: ${statement['summary']['total_expenses']:.2f}")
        print(f"  {'='*30}")
        print(f"  Net Income: ${statement['summary']['net_income']:.2f}")
        print("="*60 + "\n")
    
    def save_to_file(self, filename: str) -> None:
        """Save all transactions to a JSON file."""
        with open(filename, 'w') as f:
            json.dump([t.to_dict() for t in self.transactions], f, indent=2)
    
    def load_from_file(self, filename: str) -> None:
        """Load transactions from a JSON file."""
        with open(filename, 'r') as f:
            data = json.load(f)
            self.transactions = [Transaction.from_dict(t) for t in data]
    
    def save_to_excel(self, filename: str, start_date=None, end_date=None) -> None:
        """
        Save income statement to an Excel file.
        
        Args:
            filename: Path to save the Excel file
            start_date: Optional start date for the income statement period
            end_date: Optional end date for the income statement period
        """
        statement = self.generate_income_statement(start_date, end_date)
        
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Income Statement"
        
        # Define styles
        title_font = Font(name='Arial', size=14, bold=True)
        header_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=11)
        money_font = Font(name='Arial', size=11, bold=True)
        
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
        # Title and period
        ws['A1'] = "INCOME STATEMENT"
        ws['A1'].font = title_font
        ws.merge_cells('A1:B1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        period_text = f"Period: {statement['period']['start_date']} to {statement['period']['end_date']}"
        ws['A2'] = period_text
        ws['A2'].font = normal_font
        ws.merge_cells('A2:B2')
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Revenue section
        current_row = 4
        ws[f'A{current_row}'] = "REVENUE"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = header_fill
        ws.merge_cells(f'A{current_row}:B{current_row}')
        
        current_row += 1
        for category, amount in statement['revenue'].items():
            ws[f'A{current_row}'] = category.replace('_', ' ').title()
            ws[f'B{current_row}'] = amount
            ws[f'B{current_row}'].number_format = '$#,##0.00'
            ws[f'A{current_row}'].font = normal_font
            ws[f'B{current_row}'].font = normal_font
            current_row += 1
        
        # Total Revenue
        ws[f'A{current_row}'] = "Total Revenue"
        ws[f'B{current_row}'] = statement['summary']['total_revenue']
        ws[f'A{current_row}'].font = money_font
        ws[f'B{current_row}'].font = money_font
        ws[f'B{current_row}'].number_format = '$#,##0.00'
        ws[f'A{current_row}'].fill = total_fill
        ws[f'B{current_row}'].fill = total_fill
        
        # Expenses section
        current_row += 2
        ws[f'A{current_row}'] = "EXPENSES"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = header_fill
        ws.merge_cells(f'A{current_row}:B{current_row}')
        
        current_row += 1
        for category, amount in statement['expenses'].items():
            ws[f'A{current_row}'] = category.replace('_', ' ').title()
            ws[f'B{current_row}'] = amount
            ws[f'B{current_row}'].number_format = '$#,##0.00'
            ws[f'A{current_row}'].font = normal_font
            ws[f'B{current_row}'].font = normal_font
            current_row += 1
        
        # Total Expenses
        ws[f'A{current_row}'] = "Total Expenses"
        ws[f'B{current_row}'] = statement['summary']['total_expenses']
        ws[f'A{current_row}'].font = money_font
        ws[f'B{current_row}'].font = money_font
        ws[f'B{current_row}'].number_format = '$#,##0.00'
        ws[f'A{current_row}'].fill = total_fill
        ws[f'B{current_row}'].fill = total_fill
        
        # Summary section
        current_row += 2
        ws[f'A{current_row}'] = "SUMMARY"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = header_fill
        ws.merge_cells(f'A{current_row}:B{current_row}')
        
        current_row += 1
        ws[f'A{current_row}'] = "Total Revenue"
        ws[f'B{current_row}'] = statement['summary']['total_revenue']
        ws[f'B{current_row}'].number_format = '$#,##0.00'
        ws[f'A{current_row}'].font = normal_font
        ws[f'B{current_row}'].font = normal_font
        
        current_row += 1
        ws[f'A{current_row}'] = "Total Expenses"
        ws[f'B{current_row}'] = statement['summary']['total_expenses']
        ws[f'B{current_row}'].number_format = '$#,##0.00'
        ws[f'A{current_row}'].font = normal_font
        ws[f'B{current_row}'].font = normal_font
        
        current_row += 1
        ws[f'A{current_row}'] = "Net Income"
        ws[f'B{current_row}'] = statement['summary']['net_income']
        ws[f'A{current_row}'].font = money_font
        ws[f'B{current_row}'].font = money_font
        ws[f'B{current_row}'].number_format = '$#,##0.00'
        ws[f'A{current_row}'].fill = total_fill
        ws[f'B{current_row}'].fill = total_fill
        
        # Add borders to all cells with content
        for row in ws.iter_rows(min_row=1, max_row=current_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border
        
        # Save the workbook
        wb.save(filename)


# Example usage
if __name__ == "__main__":
    # Create income statement
    income_stmt = IncomeStatement()
    
    # Add sample transactions
    income_stmt.add_transaction("2023-01-15", "Product sales", 5000.00, "revenue", "sales")
    income_stmt.add_transaction("2023-01-20", "Consulting services", 2500.00, "revenue", "service_revenue")
    income_stmt.add_transaction("2023-01-20", "Consulting services", 2500.00, "revenue", "service_revenue")
    income_stmt.add_transaction("2023-01-20", "Consulting services", 2500.00, "revenue", "other_revenue")
    income_stmt.add_transaction("2023-01-22", "Office rent", 1200.00, "expense", "rent")
    income_stmt.add_transaction("2023-01-25", "Utility bills", 350.00, "expense", "utilities")
    income_stmt.add_transaction("2023-01-30", "Employee salaries", 3500.00, "expense", "salaries")
    
    # Print income statement
    income_stmt.print_income_statement()
    
    # Save transactions to JSON file
    income_stmt.save_to_file(r"C:\Users\shaff\OneDrive\Desktop\Autobooks_AI\transactions.json")
    
    # Save income statement to Excel file
    income_stmt.save_to_excel(r"C:\Users\shaff\OneDrive\Desktop\Autobooks_AI\income_statement.xlsx")
    
    print("Income statement generated successfully!")