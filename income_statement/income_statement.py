from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

class Transaction:
    def __init__(self, date, description, amount, category, transaction_type):
        self.date = date if isinstance(date, datetime) else datetime.strptime(date, "%Y-%m-%d")
        self.description = description
        self.amount = float(amount)
        self.category = category
        self.transaction_type = transaction_type  # 'revenue', 'expense', 'cost_of_sales', or 'inventory'

class IncomeStatement:
    def __init__(self, business_name, start_date, end_date, beginning_inventory=0):
        self.business_name = business_name
        self.start_date = start_date if isinstance(start_date, datetime) else datetime.strptime(start_date, "%Y-%m-%d")
        self.end_date = end_date if isinstance(end_date, datetime) else datetime.strptime(end_date, "%Y-%m-%d")
        self.transactions = []
        self.beginning_inventory = float(beginning_inventory)
        self.ending_inventory = None  # Will be set later
        
    def set_ending_inventory(self, ending_inventory):
        """Set the ending inventory amount"""
        self.ending_inventory = float(ending_inventory)
        
    def add_transaction(self, transaction):
        if self.start_date <= transaction.date <= self.end_date:
            self.transactions.append(transaction)
            
    def add_transactions(self, transactions):
        for transaction in transactions:
            self.add_transaction(transaction)
            
    def calculate_totals(self):
        revenue_by_category = defaultdict(float)
        expense_by_category = defaultdict(float)
        cost_of_sales_by_category = defaultdict(float)
        
        for transaction in self.transactions:
            if transaction.transaction_type == 'revenue':
                revenue_by_category[transaction.category] += transaction.amount
            elif transaction.transaction_type == 'expense':
                expense_by_category[transaction.category] += transaction.amount
            elif transaction.transaction_type == 'cost_of_sales':
                cost_of_sales_by_category[transaction.category] += transaction.amount
        
        # Calculate cost of sales with proper accounting format
        purchases = sum(cost_of_sales_by_category.values())
        total_goods_available = self.beginning_inventory + purchases
        
        if self.ending_inventory is None:
            # If ending inventory not provided, use only direct cost of sales transactions
            total_cost_of_sales = purchases
            
            # Create ordered cost of sales breakdown 
            cost_of_sales_breakdown = []
            for category, amount in cost_of_sales_by_category.items():
                cost_of_sales_breakdown.append((category, amount))
        else:
            # Use beginning inventory + purchases - ending inventory formula
            total_cost_of_sales = total_goods_available - self.ending_inventory
            
            # Create ordered cost of sales breakdown with proper accounting format
            cost_of_sales_breakdown = [
                ("Beginning Inventory", self.beginning_inventory)
            ]
            
            # Add purchases as "Plus: Goods Purchased or Manufactured"
            for category, amount in cost_of_sales_by_category.items():
                cost_of_sales_breakdown.append((category, amount))
                
            # Add total goods available
            cost_of_sales_breakdown.append(("TOTAL GOODS AVAILABLE", total_goods_available))
            
            # Add ending inventory as a deduction
            cost_of_sales_breakdown.append(("Less: Ending Inventory", -self.ending_inventory))
        
        total_revenue = sum(revenue_by_category.values())
        gross_profit = total_revenue - total_cost_of_sales
        total_expenses = sum(expense_by_category.values())
        net_income = gross_profit - total_expenses
        
        return {
            'revenue_by_category': dict(revenue_by_category),
            'cost_of_sales_breakdown': cost_of_sales_breakdown,
            'expense_by_category': dict(expense_by_category),
            'total_revenue': total_revenue,
            'total_cost_of_sales': total_cost_of_sales,
            'gross_profit': gross_profit,
            'total_expenses': total_expenses,
            'net_income': net_income
        }
    
    def generate_statement(self):
        totals = self.calculate_totals()
        
        statement = f"""
{self.business_name}
Income Statement
For the period: {self.start_date.strftime('%Y-%m-%d')} to {self.end_date.strftime('%Y-%m-%d')}

REVENUE:
"""
        
        for category, amount in totals['revenue_by_category'].items():
            statement += f"{category}: ${amount:.2f}\n"
        
        statement += f"Total Revenue: ${totals['total_revenue']:.2f}\n\n"
        
        statement += "COST OF SALES:\n"
        for category, amount in totals['cost_of_sales_breakdown']:
            statement += f"{category}: ${amount:.2f}\n"
        statement += f"Total Cost of Sales: ${totals['total_cost_of_sales']:.2f}\n\n"
        
        statement += f"GROSS PROFIT: ${totals['gross_profit']:.2f}\n\n"
        
        statement += "EXPENSES:\n"
        for category, amount in totals['expense_by_category'].items():
            statement += f"{category}: ${amount:.2f}\n"
        
        statement += f"Total Expenses: ${totals['total_expenses']:.2f}\n\n"
        statement += f"Net Income: ${totals['net_income']:.2f}\n"
        
        return statement
    
    def export_to_excel(self, filename):
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Income Statement"
        
        # Define styles
        title_font = Font(name='Arial', size=14, bold=True)
        header_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=11)
        money_font = Font(name='Arial', size=11, bold=True)
        total_font = Font(name='Arial', size=12, bold=True)
        
        # Header styling
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        center_align = Alignment(horizontal='center')
        right_align = Alignment(horizontal='right')
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        double_bottom_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='double')
        )
        
        # Helper function to apply border to a range
        def apply_border_to_range(ws, cell_range, border_style):
            start_cell, end_cell = cell_range.split(':')
            start_col, start_row = openpyxl.utils.cell.coordinate_from_string(start_cell)
            end_col, end_row = openpyxl.utils.cell.coordinate_from_string(end_cell)
            
            start_col_idx = openpyxl.utils.column_index_from_string(start_col)
            end_col_idx = openpyxl.utils.column_index_from_string(end_col)
            
            for col in range(start_col_idx, end_col_idx + 1):
                cell = ws.cell(row=int(start_row), column=col)
                cell.border = border_style
        
        # Add company name and report title
        ws['A1'] = self.business_name
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = center_align
        
        ws['A2'] = f"Income Statement"
        ws['A2'].font = header_font
        ws.merge_cells('A2:E2')
        ws['A2'].alignment = center_align
        
        ws['A3'] = f"For the period: {self.start_date.strftime('%Y-%m-%d')} to {self.end_date.strftime('%Y-%m-%d')}"
        ws['A3'].font = normal_font
        ws.merge_cells('A3:E3')
        ws['A3'].alignment = center_align
        
        # Calculate totals
        totals = self.calculate_totals()
        
        # Start row for data
        row = 5
        
        # REVENUE SECTION
        ws[f'A{row}'] = "REVENUE"
        ws[f'A{row}'].font = header_font
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        # Revenue items
        for category, amount in totals['revenue_by_category'].items():
            ws[f'B{row}'] = category
            ws[f'D{row}'] = amount
            ws[f'D{row}'].number_format = '$#,##0.00'
            row += 1
        
        # Total Revenue
        ws[f'C{row}'] = "Total Revenue"
        ws[f'C{row}'].font = money_font
        ws[f'D{row}'] = totals['total_revenue']
        ws[f'D{row}'].font = money_font
        ws[f'D{row}'].number_format = '$#,##0.00'
        apply_border_to_range(ws, f'A{row}:E{row}', double_bottom_border)
        row += 2
        
        # COST OF SALES SECTION
        ws[f'A{row}'] = "COST OF SALES"
        ws[f'A{row}'].font = header_font
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        # Cost of Sales items
        for category, amount in totals['cost_of_sales_breakdown']:
            # Make "TOTAL GOODS AVAILABLE" stand out
            if category == "TOTAL GOODS AVAILABLE":
                ws[f'B{row}'] = category
                ws[f'B{row}'].font = money_font
                ws[f'D{row}'] = amount
                ws[f'D{row}'].font = money_font
                ws[f'D{row}'].number_format = '$#,##0.00'
                apply_border_to_range(ws, f'B{row}:D{row}', thin_border)
            else:
                ws[f'B{row}'] = category
                ws[f'D{row}'] = amount
                ws[f'D{row}'].number_format = '$#,##0.00'
            row += 1
        
        # Total Cost of Sales
        ws[f'C{row}'] = "Total Cost of Sales"
        ws[f'C{row}'].font = money_font
        ws[f'D{row}'] = totals['total_cost_of_sales']
        ws[f'D{row}'].font = money_font
        ws[f'D{row}'].number_format = '$#,##0.00'
        apply_border_to_range(ws, f'A{row}:E{row}', double_bottom_border)
        row += 2
        
        # GROSS PROFIT
        ws[f'C{row}'] = "GROSS PROFIT"
        ws[f'C{row}'].font = total_font
        ws[f'D{row}'] = totals['gross_profit']
        ws[f'D{row}'].font = total_font
        ws[f'D{row}'].number_format = '$#,##0.00'
        apply_border_to_range(ws, f'A{row}:E{row}', double_bottom_border)
        row += 2
        
        # EXPENSES SECTION
        ws[f'A{row}'] = "EXPENSES"
        ws[f'A{row}'].font = header_font
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        # Expense items
        for category, amount in totals['expense_by_category'].items():
            ws[f'B{row}'] = category
            ws[f'D{row}'] = amount
            ws[f'D{row}'].number_format = '$#,##0.00'
            row += 1
        
        # Total Expenses
        ws[f'C{row}'] = "Total Expenses"
        ws[f'C{row}'].font = money_font
        ws[f'D{row}'] = totals['total_expenses']
        ws[f'D{row}'].font = money_font
        ws[f'D{row}'].number_format = '$#,##0.00'
        apply_border_to_range(ws, f'A{row}:E{row}', double_bottom_border)
        row += 2
        
        # NET INCOME
        ws[f'C{row}'] = "NET INCOME"
        ws[f'C{row}'].font = total_font
        ws[f'D{row}'] = totals['net_income']
        ws[f'D{row}'].font = total_font
        ws[f'D{row}'].number_format = '$#,##0.00'
        
        # Apply borders and format to all used cells
        for row_cells in ws.iter_rows(min_row=1, max_row=row, min_col=1, max_col=5):
            for cell in row_cells:
                if not cell.border:
                    cell.border = thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 5
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        
        # Save the workbook
        wb.save(filename)
        return filename

def main():
    # Create sample transactions
    income_statement = IncomeStatement("Sample Business LLC", "2023-01-01", "2023-03-31", beginning_inventory=2000.00)
    sample_transactions = [
        Transaction("2023-01-05", "Product Sale - Customer A", 1500.00, "Sales", "revenue"),
        Transaction("2023-01-12", "Consulting Service", 2500.00, "Services", "revenue"),
        Transaction("2023-01-08", "Raw Materials Purchase", 700.00, "Plus goods purchased or manufactured", "cost_of_sales"),
        Transaction("2023-01-10", "Inventory Purchase", 900.00, "Plus goods purchased or manufactured", "cost_of_sales"),
        Transaction("2023-01-15", "Office Rent", 800.00, "Rent", "expense"),
        Transaction("2023-01-20", "Product Sale - Customer B", 1200.00, "Sales", "revenue"),
        Transaction("2023-01-22", "Utilities", 150.00, "Utilities", "expense"),
        Transaction("2023-01-25", "Salaries", 3000.00, "Payroll", "expense"),
        Transaction("2023-02-05", "Product Sale - Customer C", 1800.00, "Sales", "revenue"),
        Transaction("2023-02-08", "Manufacturing Labor", 600.00, "Plus goods purchased or manufactured", "cost_of_sales"),
        Transaction("2023-02-10", "Marketing Campaign", 500.00, "Marketing", "expense"),
        Transaction("2023-02-15", "Office Supplies", 200.00, "Supplies", "expense"),
        Transaction("2023-02-20", "Online Course Sales", 3500.00, "Services", "revenue"),
        Transaction("2023-02-28", "Equipment Purchase", 1200.00, "Equipment", "expense"),
    ]
    
    # Create output directory if it doesn't exist
    os.makedirs("output", exist_ok=True)
    
    # Create a single income statement for Q1 2023
    income_statement.set_ending_inventory(1500.00)
    income_statement.add_transactions(sample_transactions)
    
    # Print the statement to console
    print(income_statement.generate_statement())
    
    # Export to Excel
    excel_file = income_statement.export_to_excel("output/income_statement.xlsx")
    print(f"Income statement exported to: {excel_file}")

if __name__ == "__main__":
    main()